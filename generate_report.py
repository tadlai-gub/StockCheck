import os
import datetime
import subprocess
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

# Stock holdings mapped from screenshots
STOCKS = {
    "2610.TW": {"name": "華航", "shares": 1000},
    "5271.TWO": {"name": "紘通", "shares": 2000},
    "6116.TW": {"name": "彩晶", "shares": 2000},
    "8299.TWO": {"name": "群聯", "shares": 30},
    "2002.TW": {"name": "中鋼", "shares": 1000},
    "2337.TW": {"name": "旺宏", "shares": 1000},
    "1714.TW": {"name": "和桐", "shares": 0},
}

def fetch_data():
    print("開始下載股票歷史數據...")
    # Fetch past 260 days of data (around 180 trading days) to support 100MA and Weekly 20MA calculations
    end_date = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=260)
    
    data = {}
    for ticker, info in STOCKS.items():
        try:
            print(f"下載 {info['name']} ({ticker})...")
            t = yf.Ticker(ticker)
            df = t.history(start=start_date.isoformat(), end=end_date.isoformat())
            if df.empty:
                # Fallback to general history download
                df = t.history(period="1y")
            if not df.empty:
                # Drop rows where critical fields are NaN to prevent NAType errors
                df = df.dropna(subset=['Open', 'High', 'Low', 'Close'])
            data[ticker] = df
        except Exception as e:
            print(f"下載 {ticker} 失敗: {e}")
            data[ticker] = None
    return data

def generate_candlestick_chart(ticker, df, name):
    print(f"繪製 {name} ({ticker}) 的 K線、日 100MA 與週 20MA 圖...")
    plt.rcParams['font.sans-serif'] = ['PingFang TC', 'Heiti TC', 'Arial Unicode MS', 'sans-serif']
    plt.rcParams['axes.unicode_minus'] = False
    
    df = df.copy()
    # Calculate Daily 100MA
    df['MA100'] = df['Close'].rolling(window=100).mean()
    
    # Calculate Weekly 20MA
    weekly_df = df.resample('W-FRI').last().ffill()
    weekly_df['Weekly_MA20'] = weekly_df['Close'].rolling(window=20).mean()
    
    # Map back to daily using ISO year and week
    df_yw = df.index.isocalendar()
    w_yw = weekly_df.index.isocalendar()
    df['yw'] = df_yw.year * 100 + df_yw.week
    weekly_df['yw'] = w_yw.year * 100 + w_yw.week
    ma_map = weekly_df.set_index('yw')['Weekly_MA20']
    df['Weekly_MA20'] = df['yw'].map(ma_map)
    
    # Calculate previous week's low (low of the week before the latest date's week)
    wl = df['Low'].resample('W-FRI').min()
    latest_dt = df.index[-1]
    latest_yw = latest_dt.isocalendar().year * 100 + latest_dt.isocalendar().week
    wl_before = wl[wl.index.isocalendar().year * 100 + wl.index.isocalendar().week < latest_yw]
    prev_week_low = float(wl_before.iloc[-1]) if not wl_before.empty else 0.0
    
    # Filter to last 40 trading days for plotting
    df_plot = df.tail(40)
    
    fig, ax = plt.subplots(figsize=(10, 5), dpi=100)
    
    # Background styling
    ax.set_facecolor('#F8F9FA')
    fig.patch.set_facecolor('white')
    
    # Grid lines
    ax.grid(True, which='both', linestyle='--', linewidth=0.5, color='#CCCCCC')
    
    # Taiwan stock style colors (Up is Red, Down is Green)
    up_color = '#FF3B30'   # Red
    down_color = '#34C759' # Green
    
    dates = range(len(df_plot))
    for i, (dt, row) in enumerate(df_plot.iterrows()):
        open_val = float(row['Open'])
        high_val = float(row['High'])
        low_val = float(row['Low'])
        close_val = float(row['Close'])
        
        color = up_color if close_val >= open_val else down_color
        
        # High-low wick line
        ax.plot([i, i], [low_val, high_val], color=color, linewidth=1.5)
        
        # Open-close rectangle body
        bottom = min(open_val, close_val)
        height = abs(open_val - close_val)
        if height == 0:
            height = 0.05  # Ensure flat days are visible
            
        rect = plt.Rectangle((i - 0.35, bottom), 0.7, height, facecolor=color, edgecolor=color, zorder=3)
        ax.add_patch(rect)
        
    # Plot Daily 100MA line
    ma100_vals = df_plot['MA100'].tolist()
    ax.plot(dates, ma100_vals, color='#5856D6', linewidth=1.5, label='日線 100MA', zorder=4)
    
    # Plot Weekly 20MA line (Orange)
    ma20_vals = df_plot['Weekly_MA20'].tolist()
    ax.plot(dates, ma20_vals, color='#FF9500', linewidth=2.0, label='週線 20MA (生命線)', zorder=5)
    
    # Plot Previous Week's K-line Low as Defense Line (Rose Red / Pink)
    if prev_week_low > 0:
        ax.axhline(prev_week_low, color='#FF2D55', linestyle='--', linewidth=1.5, label=f'前週K低點防守: {prev_week_low:.2f}', zorder=6)
        # Add a text label next to the defense line near the right edge
        ax.text(len(df_plot) - 1, prev_week_low, f"防守點: {prev_week_low:.2f} ", 
                color='#FF2D55', va='bottom', ha='right', fontsize=9, fontweight='bold',
                bbox=dict(facecolor='white', edgecolor='#FF2D55', boxstyle='round,pad=0.2', alpha=0.9),
                zorder=7)
    
    # Format x-axis labels (Dates)
    date_labels = [dt.strftime('%m-%d') for dt in df_plot.index]
    ax.set_xticks(dates[::5])
    ax.set_xticklabels(date_labels[::5], rotation=45, ha='right', fontsize=9)
    
    ax.set_title(f"{name} ({ticker.split('.')[0]}) K線、日 100MA 與週 20MA 走勢 (近40日)", fontsize=14, fontweight='bold', color='#1F4E78', pad=15)
    ax.set_ylabel("股價 (元)", fontsize=11)
    ax.legend(loc='upper left', frameon=True, facecolor='white', edgecolor='#CCCCCC')
    
    plt.tight_layout()
    
    img_path = f"/Users/tadlai/.gemini/antigravity/scratch/kline_{ticker.split('.')[0]}.png"
    plt.savefig(img_path, dpi=100)
    plt.close()
    return img_path

def build_excel(stock_data):
    print("開始建立 Excel 報表...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styles
    title_font = Font(name="Microsoft JhengHei", size=16, bold=True, color="1F4E78")
    subtitle_font = Font(name="Microsoft JhengHei", size=10, italic=True, color="595959")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    regular_font = Font(name="Microsoft JhengHei", size=11)
    
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    gray_fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # ------------------ Sheet 1: 投資組合總覽 (Summary) ------------------
    ws_sum = wb.create_sheet(title="投資組合總覽")
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.cell(row=2, column=2, value="個人股票投資組合與股價分析報表").font = title_font
    ws_sum.cell(row=3, column=2, value=f"報表生成日期: {datetime.date.today().strftime('%Y-%m-%d')} | 資料來源: Yahoo Finance").font = subtitle_font
    
    # Table headers
    headers = ["股票代碼", "股票名稱", "庫存股數", "最新收盤價", "週20MA(生命線)", "前週K線低點(防守)", "最新市值 (元)", "持股佔比", "防守狀態"]
    for col_idx, h in enumerate(headers, start=2):
        cell = ws_sum.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row_idx = 6
    summary_data = []
    
    for ticker, info in STOCKS.items():
        df = stock_data.get(ticker)
        if df is not None and not df.empty:
            df = df.copy()
            # Calculate weekly MA20
            weekly_df = df.resample('W-FRI').last().ffill()
            weekly_df['Weekly_MA20'] = weekly_df['Close'].rolling(window=20).mean()
            
            # Map back to daily using ISO year and week
            df_yw = df.index.isocalendar()
            w_yw = weekly_df.index.isocalendar()
            df['yw'] = df_yw.year * 100 + df_yw.week
            weekly_df['yw'] = w_yw.year * 100 + w_yw.week
            ma_map = weekly_df.set_index('yw')['Weekly_MA20']
            df['Weekly_MA20'] = df['yw'].map(ma_map)
            
            latest_row = df.iloc[-1]
            close_val = float(latest_row['Close'])
            weekly_ma20 = float(latest_row['Weekly_MA20']) if latest_row['Weekly_MA20'] == latest_row['Weekly_MA20'] else 0.0
            
            # Calculate prev week low (low of the week before the latest date's week)
            wl = df['Low'].resample('W-FRI').min()
            latest_dt = df.index[-1]
            latest_yw = latest_dt.isocalendar().year * 100 + latest_dt.isocalendar().week
            wl_before = wl[wl.index.isocalendar().year * 100 + wl.index.isocalendar().week < latest_yw]
            prev_week_low = float(wl_before.iloc[-1]) if not wl_before.empty else 0.0
        else:
            close_val, weekly_ma20, prev_week_low = 0.0, 0.0, 0.0
            
        short_code = ticker.split('.')[0]
        
        ws_sum.cell(row=row_idx, column=2, value=short_code).alignment = center_align
        ws_sum.cell(row=row_idx, column=3, value=info['name']).alignment = center_align
        ws_sum.cell(row=row_idx, column=4, value=info['shares']).number_format = '#,##0'
        ws_sum.cell(row=row_idx, column=5, value=close_val).number_format = '#,##0.00'
        
        # 週20MA (Col 6 / F)
        if weekly_ma20 > 0:
            ws_sum.cell(row=row_idx, column=6, value=weekly_ma20).number_format = '#,##0.00'
        else:
            ws_sum.cell(row=row_idx, column=6, value="-").alignment = center_align
            
        # 前週K線低點 (Col 7 / G)
        if prev_week_low > 0:
            ws_sum.cell(row=row_idx, column=7, value=prev_week_low).number_format = '#,##0.00'
        else:
            ws_sum.cell(row=row_idx, column=7, value="-").alignment = center_align
            
        # Formula for market value (Col 8 / H)
        val_formula = f"=D{row_idx}*E{row_idx}"
        ws_sum.cell(row=row_idx, column=8, value=val_formula).number_format = '#,##0.00'
        
        # Formula for percentage (Col 9 / I)
        tot_row = 6 + len(STOCKS)
        pct_formula = f"=H{row_idx}/$H${tot_row}"
        ws_sum.cell(row=row_idx, column=9, value=pct_formula).number_format = '0.0%'
        
        # Defense status (Col 10 / J)
        status_cell = ws_sum.cell(row=row_idx, column=10)
        status_cell.alignment = center_align
        if close_val == 0:
            status_cell.value = "-"
        elif close_val < prev_week_low and prev_week_low > 0:
            status_cell.value = "跌破防守"
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="9C0006")
        elif close_val < weekly_ma20 and weekly_ma20 > 0:
            status_cell.value = "跌破生命線"
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="9C6500")
        else:
            status_cell.value = "安全續抱"
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="006100")
            
        for c in range(2, 10):
            cell = ws_sum.cell(row=row_idx, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if c in [4, 5, 6, 7, 8, 9]:
                cell.alignment = right_align
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
                
        status_cell.border = thin_border
        row_idx += 1
        
    # Total row (row 13)
    tot_row = row_idx
    ws_sum.cell(row=tot_row, column=2, value="總計").alignment = center_align
    ws_sum.cell(row=tot_row, column=4, value=f"=SUM(D6:D{tot_row-1})").number_format = '#,##0'
    ws_sum.cell(row=tot_row, column=8, value=f"=SUM(H6:H{tot_row-1})").number_format = '#,##0.00'
    ws_sum.cell(row=tot_row, column=9, value=f"=SUM(I6:I{tot_row-1})").number_format = '0.0%'
    
    for c in range(2, 11):
        cell = ws_sum.cell(row=tot_row, column=c)
        cell.font = bold_font
        cell.border = total_border
        if c in [4, 5, 6, 7, 8, 9]:
            cell.alignment = right_align
            cell.alignment = right_align
            
    # Add a beautiful Pie Chart for portfolio allocation
    pie = PieChart()
    labels = Reference(ws_sum, min_col=3, min_row=6, max_row=tot_row-1)
    data = Reference(ws_sum, min_col=8, min_row=5, max_row=tot_row-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "股票持股市值分佈"
    pie.width = 16
    pie.height = 11
    ws_sum.add_chart(pie, "B14")
    
    # ------------------ Sheet 2: 兩週趨勢分析 (Two-Week Trend Analysis) ------------------
    ws_trend = wb.create_sheet(title="兩週趨勢分析")
    ws_trend.views.sheetView[0].showGridLines = True
    
    ws_trend.cell(row=2, column=2, value="近兩週股票高低價差走勢與今日焦點").font = title_font
    ws_trend.cell(row=3, column=2, value=f"區間: {(datetime.date.today() - datetime.timedelta(days=14)).strftime('%Y-%m-%d')} 至 {datetime.date.today().strftime('%Y-%m-%d')} (今日以綠色高亮)").font = subtitle_font
    
    # Highlight style for today
    today_str = datetime.date.today().strftime('%Y-%m-%d')
    highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft light green
    highlight_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="375623") # Dark green text
    
    current_row = 5
    for ticker, info in STOCKS.items():
        # Title for individual section
        ws_trend.cell(row=current_row, column=2, value=f"{info['name']} ({ticker.split('.')[0]})").font = Font(name="Microsoft JhengHei", size=14, bold=True, color="1F4E78")
        current_row += 1
        
        # Table Headers
        t_headers = ["日期", "最高價", "最低價", "收盤價"]
        for col_idx, h in enumerate(t_headers, start=2):
            cell = ws_trend.cell(row=current_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        header_row = current_row
        current_row += 1
        
        df = stock_data.get(ticker)
        data_start_row = current_row
        has_data = False
        
        if df is not None and not df.empty:
            # Filter past 14 days
            two_weeks_ago = datetime.datetime.now() - datetime.timedelta(days=14)
            df_2w = df[df.index >= two_weeks_ago.strftime('%Y-%m-%d')]
            # Sort oldest to newest (chronological for chart)
            df_2w = df_2w.sort_index(ascending=True)
            
            for dt, row in df_2w.iterrows():
                dt_str = dt.strftime('%Y-%m-%d')
                is_today = (dt_str == today_str)
                
                cell_date = ws_trend.cell(row=current_row, column=2, value=dt_str)
                cell_high = ws_trend.cell(row=current_row, column=3, value=float(row['High']))
                cell_low = ws_trend.cell(row=current_row, column=4, value=float(row['Low']))
                cell_close = ws_trend.cell(row=current_row, column=5, value=float(row['Close']))
                
                cell_date.alignment = center_align
                cell_high.number_format = '#,##0.00'
                cell_low.number_format = '#,##0.00'
                cell_close.number_format = '#,##0.00'
                
                cell_high.alignment = right_align
                cell_low.alignment = right_align
                cell_close.alignment = right_align
                
                for c in range(2, 6):
                    cell = ws_trend.cell(row=current_row, column=c)
                    cell.border = thin_border
                    if is_today:
                        cell.fill = highlight_fill
                        cell.font = highlight_font
                    else:
                        cell.font = regular_font
                        if current_row % 2 == 1:
                            cell.fill = zebra_fill
                            
                current_row += 1
                has_data = True
                
        data_end_row = current_row - 1
        
        if has_data and data_end_row >= data_start_row:
            # Create Line Chart
            chart = LineChart()
            chart.title = f"{info['name']} ({ticker.split('.')[0]}) 走勢圖"
            chart.style = 13
            chart.width = 16
            chart.height = 9.5
            
            # X-axis Categories (Dates in column 2)
            x_data = Reference(ws_trend, min_col=2, min_row=data_start_row, max_row=data_end_row)
            # Y-axis Data (High & Low in columns 3 & 4)
            y_data = Reference(ws_trend, min_col=3, max_col=4, min_row=header_row, max_row=data_end_row)
            
            chart.add_data(y_data, titles_from_data=True)
            chart.set_categories(x_data)
            
            # Style the series lines for high contrast
            if len(chart.series) >= 2:
                # Series 0 is High (最高價) -> Bright Red
                chart.series[0].graphicalProperties.line.solidFill = "FF3B30"
                chart.series[0].graphicalProperties.line.width = 25000 # 2.5pt
                # Series 1 is Low (最低價) -> Bright Blue
                chart.series[1].graphicalProperties.line.solidFill = "007AFF"
                chart.series[1].graphicalProperties.line.width = 25000 # 2.5pt
                
            chart.y_axis.title = "股價"
            chart.x_axis.number_format = 'yyyy-mm-dd'
            
            # Put the chart next to the table
            ws_trend.add_chart(chart, f"G{header_row}")
            
        # Leave spacing (20 rows from header)
        current_row = header_row + 20

    # ------------------ Sheet 3: 兩週波幅對比 (Two-Week Volatility Comparison) ------------------
    ws_vol = wb.create_sheet(title="兩週波幅對比")
    ws_vol.views.sheetView[0].showGridLines = True
    
    ws_vol.cell(row=2, column=2, value="近兩週股票每日振幅對比與今日焦點").font = title_font
    ws_vol.cell(row=3, column=2, value=f"區間: {(datetime.date.today() - datetime.timedelta(days=14)).strftime('%Y-%m-%d')} 至 {datetime.date.today().strftime('%Y-%m-%d')} (數值為每日振幅 % = (最高-最低)/最低，今日以綠色高亮)").font = subtitle_font
    
    # Headers
    vol_headers = ["日期"] + [f"{info['name']}({ticker.split('.')[0]})" for ticker, info in STOCKS.items()]
    for col_idx, h in enumerate(vol_headers, start=2):
        cell = ws_vol.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Get sorted list of dates in the last 2 weeks
    dates_2w = set()
    two_weeks_ago = datetime.datetime.now() - datetime.timedelta(days=14)
    two_weeks_ago_str = two_weeks_ago.strftime('%Y-%m-%d')
    for ticker, df in stock_data.items():
        if df is not None and not df.empty:
            df_filtered = df[df.index >= two_weeks_ago_str]
            for dt in df_filtered.index:
                dates_2w.add(dt.strftime('%Y-%m-%d'))
    sorted_dates = sorted(list(dates_2w))
    
    v_row = 6
    for dt_str in sorted_dates:
        is_today = (dt_str == today_str)
        cell_date = ws_vol.cell(row=v_row, column=2, value=dt_str)
        cell_date.alignment = center_align
        
        # Calculate amp for each stock
        for col_idx, (ticker, info) in enumerate(STOCKS.items(), start=3):
            df = stock_data.get(ticker)
            amp_val = 0.0
            if df is not None and not df.empty:
                try:
                    row = df[df.index.strftime('%Y-%m-%d') == dt_str]
                    if not row.empty:
                        high = float(row.iloc[0]['High'])
                        low = float(row.iloc[0]['Low'])
                        if low > 0:
                            amp_val = (high - low) / low
                except Exception as e:
                    pass
            cell_amp = ws_vol.cell(row=v_row, column=col_idx, value=amp_val)
            cell_amp.number_format = '0.00%'
            cell_amp.alignment = right_align
            
        for c in range(2, 2 + len(STOCKS) + 1):
            cell = ws_vol.cell(row=v_row, column=c)
            cell.border = thin_border
            if is_today:
                cell.fill = highlight_fill
                cell.font = highlight_font
            else:
                cell.font = regular_font
                if v_row % 2 == 1:
                    cell.fill = zebra_fill
        v_row += 1
        
    data_end_row = v_row - 1
    if data_end_row >= 6:
        chart = LineChart()
        chart.title = "近兩週每日振幅對比走勢圖"
        chart.style = 10
        chart.width = 18
        chart.height = 12
        
        x_data = Reference(ws_vol, min_col=2, min_row=6, max_row=data_end_row)
        y_data = Reference(ws_vol, min_col=3, max_col=2 + len(STOCKS), min_row=5, max_row=data_end_row)
        
        chart.add_data(y_data, titles_from_data=True)
        chart.set_categories(x_data)
        chart.y_axis.title = "振幅 %"
        chart.x_axis.number_format = 'yyyy-mm-dd'
        
        ws_vol.add_chart(chart, f"B{data_end_row + 3}")

    # ------------------ Sheet 4: 合併歷史數據 (Combined History) ------------------
    ws_all = wb.create_sheet(title="合併歷史數據")
    ws_all.views.sheetView[0].showGridLines = True
    ws_all.freeze_panes = "A2"
    
    all_headers = ["交易日期", "股票代碼", "股票名稱", "開盤價", "最高價", "最低價", "收盤價", "每日高低差", "高低振幅", "成交量"]
    for col_idx, h in enumerate(all_headers, start=1):
        cell = ws_all.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = gray_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    all_rows = []
    for ticker, info in STOCKS.items():
        df = stock_data.get(ticker)
        if df is not None and not df.empty:
            for dt, row in df.iterrows():
                dt_str = dt.strftime('%Y-%m-%d')
                all_rows.append({
                    "Date": dt_str,
                    "Code": ticker.split('.')[0],
                    "Name": info['name'],
                    "Open": float(row['Open']),
                    "High": float(row['High']),
                    "Low": float(row['Low']),
                    "Close": float(row['Close']),
                    "Volume": int(row['Volume'])
                })
                
    # Sort by Date desc, Code asc
    all_rows.sort(key=lambda x: (x["Date"], x["Code"]), reverse=True)
    
    for r_idx, r in enumerate(all_rows, start=2):
        ws_all.cell(row=r_idx, column=1, value=r["Date"]).alignment = center_align
        ws_all.cell(row=r_idx, column=2, value=r["Code"]).alignment = center_align
        ws_all.cell(row=r_idx, column=3, value=r["Name"]).alignment = center_align
        ws_all.cell(row=r_idx, column=4, value=r["Open"]).number_format = '#,##0.00'
        ws_all.cell(row=r_idx, column=5, value=r["High"]).number_format = '#,##0.00'
        ws_all.cell(row=r_idx, column=6, value=r["Low"]).number_format = '#,##0.00'
        ws_all.cell(row=r_idx, column=7, value=r["Close"]).number_format = '#,##0.00'
        
        # Formulas
        ws_all.cell(row=r_idx, column=8, value=f"=E{r_idx}-F{r_idx}").number_format = '#,##0.00'
        ws_all.cell(row=r_idx, column=9, value=f"=H{r_idx}/F{r_idx}").number_format = '0.00%'
        ws_all.cell(row=r_idx, column=10, value=r["Volume"]).number_format = '#,##0'
        
        for c in range(1, 11):
            cell = ws_all.cell(row=r_idx, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if c in [4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = right_align
            if r_idx % 2 == 1:
                cell.fill = zebra_fill
                
    # ------------------ Sheets 3-8: Individual Stock Sheets ------------------
    for ticker, info in STOCKS.items():
        sheet_title = f"{info['name']}({ticker.split('.')[0]})"
        ws_stock = wb.create_sheet(title=sheet_title)
        ws_stock.views.sheetView[0].showGridLines = True
        ws_stock.freeze_panes = "A5"
        
        ws_stock.cell(row=2, column=1, value=f"{info['name']} ({ticker}) 歷史股價高低分析日報表").font = title_font
        
        df = stock_data.get(ticker)
        prev_week_low = 0.0
        current_weekly_20ma = 0.0
        close_val = 0.0
        if df is not None and not df.empty:
            try:
                close_val = float(df['Close'].iloc[-1])
                # Calculate weekly low (low of the week before the latest date's week)
                wl = df['Low'].resample('W-FRI').min()
                latest_dt = df.index[-1]
                latest_yw = latest_dt.isocalendar().year * 100 + latest_dt.isocalendar().week
                wl_before = wl[wl.index.isocalendar().year * 100 + wl.index.isocalendar().week < latest_yw]
                prev_week_low = float(wl_before.iloc[-1]) if not wl_before.empty else 0.0
                
                # Calculate weekly 20ma
                weekly_df = df.resample('W-FRI').last().ffill()
                weekly_df['Weekly_MA20'] = weekly_df['Close'].rolling(window=20).mean()
                current_weekly_20ma = float(weekly_df['Weekly_MA20'].iloc[-1]) if not weekly_df.empty and weekly_df['Weekly_MA20'].iloc[-1] == weekly_df['Weekly_MA20'].iloc[-1] else 0.0
            except Exception:
                pass
                
        ws_stock.cell(row=3, column=1, value=f"庫存股數: {info['shares']} 股").font = bold_font
        
        # Display weekly 20MA
        ma20_cell = ws_stock.cell(row=3, column=4, value=f"目前週20MA (生命線): {f'{current_weekly_20ma:.2f} 元' if current_weekly_20ma > 0 else 'N/A'}")
        ma20_cell.font = bold_font
        
        # Display defense price
        defense_cell = ws_stock.cell(row=3, column=7, value=f"前週K線低點 (防守點): {f'{prev_week_low:.2f} 元' if prev_week_low > 0 else 'N/A'}")
        defense_cell.font = bold_font
        
        # Highlight and warn if broken defense or MA20
        if close_val < prev_week_low and prev_week_low > 0:
            defense_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            defense_cell.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="9C0006")
            ws_stock.cell(row=3, column=10, value="⚠️ 跌破防守點！").font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FF3B30")
        elif close_val < current_weekly_20ma and current_weekly_20ma > 0:
            ma20_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            ma20_cell.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="9C6500")
            ws_stock.cell(row=3, column=10, value="⚠️ 跌破生命線！").font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FF9500")
        
        stock_headers = ["交易日期", "開盤價", "最高價", "最低價", "收盤價", "每日高低差", "高低振幅", "成交量", "日100MA", "週20MA(生命線)"]
        for col_idx, h in enumerate(stock_headers, start=1):
            cell = ws_stock.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        if df is not None and not df.empty:
            df = df.copy()
            # Calculate Daily 100MA
            df['MA100'] = df['Close'].rolling(window=100).mean()
            
            # Calculate Weekly 20MA
            weekly_df = df.resample('W-FRI').last().ffill()
            weekly_df['Weekly_MA20'] = weekly_df['Close'].rolling(window=20).mean()
            df_yw = df.index.isocalendar()
            w_yw = weekly_df.index.isocalendar()
            df['yw'] = df_yw.year * 100 + df_yw.week
            weekly_df['yw'] = w_yw.year * 100 + w_yw.week
            ma_map = weekly_df.set_index('yw')['Weekly_MA20']
            df['Weekly_MA20'] = df['yw'].map(ma_map)
            
            # Generate the K-line chart image
            try:
                img_path = generate_candlestick_chart(ticker, df, info['name'])
                img = Image(img_path)
                # Place K-line chart starting from column L (Column 12), row 2
                ws_stock.add_image(img, "L2")
            except Exception as e:
                print(f"繪製 {ticker} K線圖失敗: {e}")
                
            # Sort descending by date
            df_sorted = df.sort_index(ascending=False)
            for r_idx, (dt, row) in enumerate(df_sorted.iterrows(), start=5):
                dt_str = dt.strftime('%Y-%m-%d')
                ws_stock.cell(row=r_idx, column=1, value=dt_str).alignment = center_align
                ws_stock.cell(row=r_idx, column=2, value=float(row['Open'])).number_format = '#,##0.00'
                ws_stock.cell(row=r_idx, column=3, value=float(row['High'])).number_format = '#,##0.00'
                ws_stock.cell(row=r_idx, column=4, value=float(row['Low'])).number_format = '#,##0.00'
                ws_stock.cell(row=r_idx, column=5, value=float(row['Close'])).number_format = '#,##0.00'
                
                # Formulas
                ws_stock.cell(row=r_idx, column=6, value=f"=C{r_idx}-D{r_idx}").number_format = '#,##0.00'
                ws_stock.cell(row=r_idx, column=7, value=f"=F{r_idx}/D{r_idx}").number_format = '0.00%'
                ws_stock.cell(row=r_idx, column=8, value=int(row['Volume'])).number_format = '#,##0'
                
                # MA100 value (Col 9)
                ma100_val = row['MA100']
                if ma100_val == ma100_val:  # Check if not NaN
                    ws_stock.cell(row=r_idx, column=9, value=float(ma100_val)).number_format = '#,##0.00'
                else:
                    ws_stock.cell(row=r_idx, column=9, value="-").alignment = center_align
                    
                # Weekly MA20 value (Col 10)
                ma20_val = row['Weekly_MA20']
                if ma20_val == ma20_val:  # Check if not NaN
                    ws_stock.cell(row=r_idx, column=10, value=float(ma20_val)).number_format = '#,##0.00'
                else:
                    ws_stock.cell(row=r_idx, column=10, value="-").alignment = center_align
                
                for c in range(1, 11):
                    cell = ws_stock.cell(row=r_idx, column=c)
                    cell.font = regular_font
                    cell.border = thin_border
                    if c in [2, 3, 4, 5, 6, 7, 8, 9, 10]:
                        cell.alignment = right_align
                    if r_idx % 2 == 1:
                        cell.fill = zebra_fill
                        
    # Auto-fit columns across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Exclude title rows in auto-fit calculation to avoid huge columns
                if cell.row in [2, 3] and sheet.title in ["投資組合總覽"] or (cell.row in [2, 3] and sheet.title not in ["投資組合總覽", "合併歷史數據"]):
                    continue
                val = str(cell.value or '')
                if cell.number_format and '%' in cell.number_format:
                    val += '%%'
                max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save the file
    docs_path = "/Users/tadlai/Documents/Work/Antigravity/股票高低價差日報表.xlsx"
    desktop_path = "/Users/tadlai/Desktop/股票高低價差日報表.xlsx"
    
    # Delete Desktop copy if exists
    if os.path.exists(desktop_path):
        try:
            os.remove(desktop_path)
            print(f"已刪除桌面備份: {desktop_path}")
        except Exception as e:
            print(f"無法刪除桌面備份 {desktop_path}: {e}")
            
    # Delete old Documents root copy if exists
    old_docs_path = "/Users/tadlai/Documents/股票高低價差日報表.xlsx"
    if os.path.exists(old_docs_path):
        try:
            os.remove(old_docs_path)
            print(f"已刪除舊檔案: {old_docs_path}")
        except Exception as e:
            print(f"無法刪除舊檔案 {old_docs_path}: {e}")
            
    # Ensure directory exists
    os.makedirs(os.path.dirname(docs_path), exist_ok=True)
    
    wb.save(docs_path)
    print(f"Excel 報表已儲存至: {docs_path}")
    return docs_path

def send_email(file_path, recipient_email):
    print(f"嘗試發送電子郵件至 {recipient_email}...")
    subject = "股票股價高低區間日報表"
    body = "Tad,\\n\\n您好！附件為您所關注股票（華航、紘通、彩晶、群聯、中鋼、旺宏、和桐）的股價高低區間分析日報表。\\n\\n報表包含：\\n1. 投資組合總覽（含最新市值及佔比餅圖）\\n2. 合併歷史日數據（依日期排序）\\n3. 個股歷史高低價差及振幅分析頁籤\\n\\n祝您投資順利！\\n\\n您的 AI 助理 Antigravity"
    
    # Method 1: AppleScript via osascript to trigger Mail.app (the standard macOS way)
    applescript = f'''
    tell application "Mail"
        set newEmail to make new outgoing message with properties {{subject:"{subject}", content:"{body}", visible:true}}
        tell newEmail
            make new to recipient at end of to recipients with properties {{address:"{recipient_email}"}}
            make new attachment with properties {{file name:(POSIX file "{file_path}")}} at after the last paragraph
            send
        end tell
    end tell
    '''
    
    # Save AppleScript to temp file
    as_file = "/Users/tadlai/.gemini/antigravity/scratch/send_mail.scpt"
    with open(as_file, 'w', encoding='utf-8') as f:
        f.write(applescript)
        
    try:
        print("執行 AppleScript 以透過 Mail.app 發送郵件...")
        result = subprocess.run(["osascript", as_file], capture_output=True, text=True, check=True)
        print("AppleScript 發送指令執行成功！")
        return True
    except Exception as e:
        print(f"AppleScript 發送失敗: {e}")
        
    # Method 2: Try standard CLI mail or sendmail (for local SMTP daemon if running)
    print("AppleScript 未成功或受阻，嘗試使用 local sendmail...")
    try:
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        
        msg = MIMEMultipart()
        msg['From'] = "tad.lai@icloud.com"
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body.replace("\\n", "\n"), 'plain', 'utf-8'))
        
        with open(file_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {os.path.basename(file_path)}",
            )
            msg.attach(part)
            
        sendmail_proc = subprocess.Popen(["/usr/sbin/sendmail", "-t", "-oi"], stdin=subprocess.PIPE, text=True)
        sendmail_proc.communicate(msg.as_string())
        print("Sendmail 指令已執行。")
        return True
    except Exception as e:
        print(f"Sendmail 發送失敗: {e}")
        
    return False

if __name__ == "__main__":
    stock_data = fetch_data()
    file_path = build_excel(stock_data)
    success = send_email(file_path, "tad.lai@icloud.com")
    if success:
        print("電子郵件發送作業已啟動！")
    else:
        print("郵件自動發送可能未完全成功，請手動確認或在桌面查看檔案。")
